import openpyxl
book = openpyxl.load_workbook("E:\\ADITYA\\Selenium\\pythonSelfFramework\\PythonDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet["A5"].value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value == "TestCase2":
        for j in range(1,sheet.max_column+1):
            print(sheet.cell(row=i,column=j).value)