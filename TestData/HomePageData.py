import openpyxl


class HomePageData:

    test_HomePage_Data = [{"firstname":"Aditya", "email":"akstr17official@gmail.com","password":"1234", "gender":"Male"},
                          {"firstname":"Striker", "email":"akstr17obussiness@gmail.com", "password":"1234", "gender":"Male"},
                          {"firstname":"Toshaka", "email":"akstr17gaming@gmail.com", "password":"1234", "gender":"Female"}]
    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("E:\\ADITYA\\Selenium\\pythonSelfFramework\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return[Dict]

